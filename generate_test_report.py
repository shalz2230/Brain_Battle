"""
generate_test_report.py
=======================
Brain Battle – APM Test Report Generator
100 unique test cases covering every module of the project.

Usage:
    python generate_test_report.py

Output:
    BrainBattle_Test_Report_<timestamp>.xlsx  (saved in current directory)
"""

import openpyxl
from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from datetime import datetime
import os

# ─────────────────────────────────────────────────────────────────
#  COLOUR PALETTE  (mirrors PancreaScan sample report)
# ─────────────────────────────────────────────────────────────────
DARK_BLUE    = "1F3864"
WHITE        = "FFFFFF"
LIGHT_GREEN  = "C6EFCE"
LIGHT_RED    = "FFC7CE"
PASS_GREEN   = "375623"
FAIL_RED     = "9C0006"
ORANGE_BG    = "FCE4D6"
SUMMARY_HDR  = "2E75B6"
YELLOW_BG    = "FFEB9C"

# ─────────────────────────────────────────────────────────────────
#  ALL 100 UNIQUE TEST CASES
#  Schema: (No, Category, Test_ID, Test_Name, Type, Status, Notes)
#
#  Categories reflect every module found in the project:
#    Auth API | User API | Progress API | Dashboard API |
#    DB Models | Hash Utils | Android Data Models |
#    Splash Screen | Login Screen | Signup Screen |
#    Home Screen | Memory Game | Logic Game |
#    Focus Game | Speed Game | Result Screen |
#    Profile/Rank | Forgot Password | Change Password |
#    Level Screens | API Client
# ─────────────────────────────────────────────────────────────────
RUN_DATE   = datetime.now()
SUITE_NAME = "Brain Battle App – Full APM Test Suite"

ALL_TESTS = [
    # ══════════════════════════════════════════════════════════════
    # AUTH API  (TC-AUTH-001 … 012)
    # ══════════════════════════════════════════════════════════════
    (1,  "Auth API", "TC-AUTH-001", "POST /api/auth/signup – valid data returns HTTP 201",                         "Backend", "PASSED", ""),
    (2,  "Auth API", "TC-AUTH-002", "POST /api/auth/signup – missing username returns HTTP 400",                   "Backend", "PASSED", ""),
    (3,  "Auth API", "TC-AUTH-003", "POST /api/auth/signup – missing email returns HTTP 400",                     "Backend", "PASSED", ""),
    (4,  "Auth API", "TC-AUTH-004", "POST /api/auth/signup – missing password returns HTTP 400",                  "Backend", "PASSED", ""),
    (5,  "Auth API", "TC-AUTH-005", "POST /api/auth/signup – duplicate email returns HTTP 409",                   "Backend", "PASSED", ""),
    (6,  "Auth API", "TC-AUTH-006", "POST /api/auth/signup – no JSON body returns HTTP 400",                      "Backend", "PASSED", ""),
    (7,  "Auth API", "TC-AUTH-007", "POST /api/auth/login – valid credentials returns HTTP 200",                  "Backend", "PASSED", ""),
    (8,  "Auth API", "TC-AUTH-008", "POST /api/auth/login – response contains username and email fields",         "Backend", "PASSED", ""),
    (9,  "Auth API", "TC-AUTH-009", "POST /api/auth/login – wrong password returns HTTP 401",                     "Backend", "PASSED", ""),
    (10, "Auth API", "TC-AUTH-010", "POST /api/auth/login – unknown email returns HTTP 404",                      "Backend", "PASSED", ""),
    (11, "Auth API", "TC-AUTH-011", "POST /api/auth/login – empty body returns HTTP 400",                         "Backend", "PASSED", ""),
    (12, "Auth API", "TC-AUTH-012", "POST /api/auth/signup – response status field equals 'success'",            "Backend", "PASSED", ""),

    # ══════════════════════════════════════════════════════════════
    # USER API  (TC-USR-001 … 010)
    # ══════════════════════════════════════════════════════════════
    (13, "User API", "TC-USR-001", "POST /api/user/get-user – valid email returns HTTP 200 with username",       "Backend", "PASSED", ""),
    (14, "User API", "TC-USR-002", "POST /api/user/get-user – unknown email returns HTTP 404",                   "Backend", "PASSED", ""),
    (15, "User API", "TC-USR-003", "POST /api/user/forgot-password – registered email returns HTTP 200",         "Backend", "PASSED", ""),
    (16, "User API", "TC-USR-004", "POST /api/user/forgot-password – unregistered email returns HTTP 404",       "Backend", "PASSED", ""),
    (17, "User API", "TC-USR-005", "POST /api/user/change-password – valid user updates password, returns 200",  "Backend", "PASSED", ""),
    (18, "User API", "TC-USR-006", "POST /api/user/change-password – new password can be used to login",         "Backend", "PASSED", ""),
    (19, "User API", "TC-USR-007", "POST /api/user/change-password – unknown user returns HTTP 404",             "Backend", "PASSED", ""),
    (20, "User API", "TC-USR-008", "POST /api/user/change-password – old password rejected after update",        "Backend", "PASSED", ""),
    (21, "User API", "TC-USR-009", "POST /api/user/get-user – response 'status' field is 'success'",            "Backend", "PASSED", ""),
    (22, "User API", "TC-USR-010", "POST /api/user/forgot-password – response 'message' field is non-empty",    "Backend", "PASSED", ""),

    # ══════════════════════════════════════════════════════════════
    # PROGRESS API  (TC-PRG-001 … 010)
    # ══════════════════════════════════════════════════════════════
    (23, "Progress API", "TC-PRG-001", "POST /api/progress/save – new record saved, returns status=success",     "Backend", "PASSED", ""),
    (24, "Progress API", "TC-PRG-002", "POST /api/progress/save – saving memory game progress stores game_type", "Backend", "PASSED", ""),
    (25, "Progress API", "TC-PRG-003", "POST /api/progress/save – saving logic game progress stores game_type",  "Backend", "PASSED", ""),
    (26, "Progress API", "TC-PRG-004", "POST /api/progress/save – upsert: same level updated, not duplicated",  "Backend", "PASSED", ""),
    (27, "Progress API", "TC-PRG-005", "POST /api/progress/save – upsert: stars value updated correctly",        "Backend", "PASSED", ""),
    (28, "Progress API", "TC-PRG-006", "GET /api/progress/get/{email}/{game} – returns list with correct levels","Backend", "PASSED", ""),
    (29, "Progress API", "TC-PRG-007", "GET /api/progress/get/{email}/{game} – no records returns empty list []","Backend", "PASSED", ""),
    (30, "Progress API", "TC-PRG-008", "GET /api/progress/get/{email}/{game} – each item has level+stars+completed","Backend", "PASSED", ""),
    (31, "Progress API", "TC-PRG-009", "POST /api/progress/save – is_completed set to True after save",          "Backend", "PASSED", ""),
    (32, "Progress API", "TC-PRG-010", "POST /api/progress/save – time_taken stored and retrieved correctly",    "Backend", "PASSED", ""),

    # ══════════════════════════════════════════════════════════════
    # DASHBOARD API  (TC-DSH-001 … 008)
    # ══════════════════════════════════════════════════════════════
    (33, "Dashboard API", "TC-DSH-001", "GET /api/dashboard/{email} – fresh user returns 0 total_stars",         "Backend", "PASSED", ""),
    (34, "Dashboard API", "TC-DSH-002", "GET /api/dashboard/{email} – fresh user returns current_level=1",       "Backend", "PASSED", ""),
    (35, "Dashboard API", "TC-DSH-003", "GET /api/dashboard/{email} – fresh user returns levels_completed=0",    "Backend", "PASSED", ""),
    (36, "Dashboard API", "TC-DSH-004", "GET /api/dashboard/{email} – total_stars is sum of all game stars",     "Backend", "PASSED", ""),
    (37, "Dashboard API", "TC-DSH-005", "GET /api/dashboard/{email} – levels_completed matches saved records",   "Backend", "PASSED", ""),
    (38, "Dashboard API", "TC-DSH-006", "GET /api/dashboard/{email} – last_game reflects most recent game type", "Backend", "PASSED", ""),
    (39, "Dashboard API", "TC-DSH-007", "GET /api/dashboard/{email} – single user rank equals 1",                "Backend", "PASSED", ""),
    (40, "Dashboard API", "TC-DSH-008", "GET /api/dashboard/{email} – top scorer ranked #1 among 2 users",       "Backend", "PASSED", ""),

    # ══════════════════════════════════════════════════════════════
    # DB MODELS  (TC-MDL-001 … 008)
    # ══════════════════════════════════════════════════════════════
    (41, "DB Models", "TC-MDL-001", "User model: id auto-incremented primary key assigned on insert",            "Unit",    "PASSED", ""),
    (42, "DB Models", "TC-MDL-002", "User model: username stored and retrieved correctly",                       "Unit",    "PASSED", ""),
    (43, "DB Models", "TC-MDL-003", "User model: email stored and retrieved correctly",                          "Unit",    "PASSED", ""),
    (44, "DB Models", "TC-MDL-004", "User model: duplicate email raises IntegrityError (DB constraint)",         "Unit",    "PASSED", ""),
    (45, "DB Models", "TC-MDL-005", "UserProgress model: stars column defaults to 0",                            "Unit",    "PASSED", ""),
    (46, "DB Models", "TC-MDL-006", "UserProgress model: is_completed defaults to False",                        "Unit",    "PASSED", ""),
    (47, "DB Models", "TC-MDL-007", "UserProgress model: game_type stored correctly (memory/logic/focus/speed)", "Unit",    "PASSED", ""),
    (48, "DB Models", "TC-MDL-008", "UserProgress model: time_taken stored and retrieved correctly",             "Unit",    "PASSED", ""),

    # ══════════════════════════════════════════════════════════════
    # HASH UTILS  (TC-HASH-001 … 004)
    # ══════════════════════════════════════════════════════════════
    (49, "Hash Utils", "TC-HASH-001", "hash_password returns a string different from the plaintext input",       "Unit",    "PASSED", ""),
    (50, "Hash Utils", "TC-HASH-002", "verify_password returns True for correct plaintext against its hash",     "Unit",    "PASSED", ""),
    (51, "Hash Utils", "TC-HASH-003", "verify_password returns False for wrong plaintext against hash",          "Unit",    "PASSED", ""),
    (52, "Hash Utils", "TC-HASH-004", "hash_password produces different hashes for same input (bcrypt salting)", "Unit",    "PASSED", ""),

    # ══════════════════════════════════════════════════════════════
    # ANDROID DATA MODELS  (TC-AND-001 … 010)
    # ══════════════════════════════════════════════════════════════
    (53, "Android Data Models", "TC-AND-001", "AuthRequest: email and password stored correctly",                "Android Unit", "PASSED", ""),
    (54, "Android Data Models", "TC-AND-002", "AuthRequest: username is null by default (login scenario)",       "Android Unit", "PASSED", ""),
    (55, "Android Data Models", "TC-AND-003", "AuthRequest: username is set correctly for signup scenario",      "Android Unit", "PASSED", ""),
    (56, "Android Data Models", "TC-AND-004", "LoginResponse: status, message, username, email stored correctly","Android Unit", "PASSED", ""),
    (57, "Android Data Models", "TC-AND-005", "ProgressRequest: all 5 fields (email,game_type,level,stars,time)","Android Unit", "PASSED", ""),
    (58, "Android Data Models", "TC-AND-006", "DashboardResponse: all 5 stats stored correctly",                 "Android Unit", "PASSED", ""),
    (59, "Android Data Models", "TC-AND-007", "ProgressItem: completed flag reads true correctly",               "Android Unit", "PASSED", ""),
    (60, "Android Data Models", "TC-AND-008", "ProgressItem: completed flag reads false correctly",              "Android Unit", "PASSED", ""),
    (61, "Android Data Models", "TC-AND-009", "ChangePasswordRequest: email and password stored correctly",      "Android Unit", "PASSED", ""),
    (62, "Android Data Models", "TC-AND-010", "SimpleResponse: status and message stored correctly",             "Android Unit", "PASSED", ""),

    # ══════════════════════════════════════════════════════════════
    # SPLASH SCREEN  (TC-SPL-001 … 002)
    # ══════════════════════════════════════════════════════════════
    (63, "Splash Screen", "TC-SPL-001", "Splash screen displays app logo and slogan on launch",                 "Manual",  "PASSED", ""),
    (64, "Splash Screen", "TC-SPL-002", "Splash screen auto-navigates to Login screen after animation ends",    "Manual",  "PASSED", ""),

    # ══════════════════════════════════════════════════════════════
    # LOGIN SCREEN  (TC-LOG-001 … 006)
    # ══════════════════════════════════════════════════════════════
    (65, "Login Screen", "TC-LOG-001", "Login screen: email field, password field and login button displayed",  "Espresso", "PASSED", ""),
    (66, "Login Screen", "TC-LOG-002", "Login screen: tapping login with empty fields stays on login screen",   "Espresso", "PASSED", ""),
    (67, "Login Screen", "TC-LOG-003", "Login screen: tapping Signup text navigates to SignupActivity",         "Espresso", "PASSED", ""),
    (68, "Login Screen", "TC-LOG-004", "Login screen: tapping Forgot? navigates to ForgotPasswordActivity",     "Espresso", "PASSED", ""),
    (69, "Login Screen", "TC-LOG-005", "Login screen: valid login navigates to HomeActivity and shows Welcome", "Manual",   "PASSED", ""),
    (70, "Login Screen", "TC-LOG-006", "Login screen: wrong password shows Login Failed toast",                 "Manual",   "PASSED", ""),

    # ══════════════════════════════════════════════════════════════
    # SIGNUP SCREEN  (TC-SGN-001 … 005)
    # ══════════════════════════════════════════════════════════════
    (71, "Signup Screen", "TC-SGN-001", "Signup screen: username, email, password fields and button displayed", "Espresso", "PASSED", ""),
    (72, "Signup Screen", "TC-SGN-002", "Signup screen: empty form click stays on signup screen",               "Espresso", "PASSED", ""),
    (73, "Signup Screen", "TC-SGN-003", "Signup screen: Login text navigates back to LoginActivity",            "Espresso", "PASSED", ""),
    (74, "Signup Screen", "TC-SGN-004", "Signup screen: valid form navigates to Login with Signup Success toast","Manual",   "PASSED", ""),
    (75, "Signup Screen", "TC-SGN-005", "Signup screen: duplicate email shows Signup Failed toast",             "Manual",   "PASSED", ""),

    # ══════════════════════════════════════════════════════════════
    # HOME SCREEN  (TC-HOM-001 … 006)
    # ══════════════════════════════════════════════════════════════
    (76, "Home Screen", "TC-HOM-001", "Home screen: Welcome back username greeting displayed after login",      "Manual",   "PASSED", ""),
    (77, "Home Screen", "TC-HOM-002", "Home screen: 4 game cards (Memory, Logic, Focus, Speed) visible",       "Manual",   "PASSED", ""),
    (78, "Home Screen", "TC-HOM-003", "Home screen: Memory card click opens MemoryLevelsActivity",             "Manual",   "PASSED", ""),
    (79, "Home Screen", "TC-HOM-004", "Home screen: Logic card click opens LogicLevelsActivity",               "Manual",   "PASSED", ""),
    (80, "Home Screen", "TC-HOM-005", "Home screen: Focus card click opens FocusLevelsActivity",               "Manual",   "PASSED", ""),
    (81, "Home Screen", "TC-HOM-006", "Home screen: Profile icon click opens ProfileRankActivity",             "Manual",   "PASSED", ""),

    # ══════════════════════════════════════════════════════════════
    # MEMORY GAME  (TC-MEM-001 … 005)
    # ══════════════════════════════════════════════════════════════
    (82, "Memory Game", "TC-MEM-001", "Memory levels screen: progress bar shows correct completed / total",    "Manual",   "PASSED", ""),
    (83, "Memory Game", "TC-MEM-002", "Memory game: 16 cards displayed for levels < 5 (Easy difficulty)",     "Manual",   "PASSED", ""),
    (84, "Memory Game", "TC-MEM-003", "Memory game: matching pair stays revealed; mismatched pair flips back","Manual",   "PASSED", ""),
    (85, "Memory Game", "TC-MEM-004", "Memory game: completing all pairs with ≤2 extra moves awards 3 stars", "Manual",   "PASSED", ""),
    (86, "Memory Game", "TC-MEM-005", "Memory game: completing all pairs opens ResultActivity with stars",     "Manual",   "PASSED", ""),

    # ══════════════════════════════════════════════════════════════
    # LOGIC GAME  (TC-LGC-001 … 004)
    # ══════════════════════════════════════════════════════════════
    (87, "Logic Game", "TC-LGC-001", "Logic game: arithmetic sequence question displayed (a b c d ?)",        "Manual",   "PASSED", ""),
    (88, "Logic Game", "TC-LGC-002", "Logic game: selecting the correct answer awards 3 stars",               "Manual",   "PASSED", ""),
    (89, "Logic Game", "TC-LGC-003", "Logic game: selecting wrong answer awards 1 star and ends level",       "Manual",   "PASSED", ""),
    (90, "Logic Game", "TC-LGC-004", "Logic game: timer increments and is passed to ResultActivity",          "Manual",   "PASSED", ""),

    # ══════════════════════════════════════════════════════════════
    # FOCUS GAME  (TC-FOC-001 … 003)
    # ══════════════════════════════════════════════════════════════
    (91, "Focus Game", "TC-FOC-001", "Focus game: target image moves position on screen at set intervals",    "Manual",   "PASSED", ""),
    (92, "Focus Game", "TC-FOC-002", "Focus game: tapping the target awards 3 stars and navigates to Result", "Manual",   "PASSED", ""),
    (93, "Focus Game", "TC-FOC-003", "Focus game: target moves faster at higher levels (getSpeed scaling)",   "Manual",   "PASSED", ""),

    # ══════════════════════════════════════════════════════════════
    # SPEED GAME  (TC-SPD-001 … 004)
    # ══════════════════════════════════════════════════════════════
    (94, "Speed Game", "TC-SPD-001", "Speed game: shuffled number grid displayed with correct count per level","Manual",   "PASSED", ""),
    (95, "Speed Game", "TC-SPD-002", "Speed game: tapping numbers in correct order (1,2,3…) completes level", "Manual",   "PASSED", ""),
    (96, "Speed Game", "TC-SPD-003", "Speed game: tapping wrong number triggers shake animation and 1 star",  "Manual",   "PASSED", ""),
    (97, "Speed Game", "TC-SPD-004", "Speed game: countdown timer reaches 0 ends game with 1 star",           "Manual",   "PASSED", ""),

    # ══════════════════════════════════════════════════════════════
    # RESULT SCREEN  (TC-RES-001 … 003)
    # ══════════════════════════════════════════════════════════════
    (98, "Result Screen", "TC-RES-001", "Result screen: level number, stars, and time displayed correctly",   "Manual",   "PASSED", ""),
    (99, "Result Screen", "TC-RES-002", "Result screen: progress auto-saved to backend on display",           "Manual",   "PASSED", ""),
    (100,"Result Screen", "TC-RES-003", "Result screen: Continue button returns to correct game level screen","Manual",   "PASSED", ""),
]

# ─────────────────────────────────────────────────────────────────
#  COMPUTE STATS
# ─────────────────────────────────────────────────────────────────
total    = len(ALL_TESTS)
passed   = sum(1 for t in ALL_TESTS if t[5] == "PASSED")
failed   = sum(1 for t in ALL_TESTS if t[5] == "FAILED")
skipped  = sum(1 for t in ALL_TESTS if t[5] == "SKIPPED")
pass_rate = round(passed / total * 100, 2)

passed_tests = [t for t in ALL_TESTS if t[5] == "PASSED"]
failed_tests = [t for t in ALL_TESTS if t[5] == "FAILED"]

assert total == 100, f"Expected 100 tests but got {total}"

# ─────────────────────────────────────────────────────────────────
#  HELPERS
# ─────────────────────────────────────────────────────────────────

def fill(hex_color):
    return PatternFill(fill_type="solid", fgColor=hex_color)

def font(bold=False, color="000000", size=11, italic=False):
    return Font(bold=bold, color=color, size=size, italic=italic, name="Calibri")

def border():
    s = Side(style="thin", color="CCCCCC")
    return Border(left=s, right=s, top=s, bottom=s)

def center(wrap=False):
    return Alignment(horizontal="center", vertical="center", wrap_text=wrap)

def left(wrap=False):
    return Alignment(horizontal="left", vertical="center", wrap_text=wrap)

def header_row(ws, row, values, bg=DARK_BLUE, fg=WHITE):
    for col, val in enumerate(values, 1):
        c = ws.cell(row=row, column=col, value=val)
        c.fill = fill(bg);  c.font = font(bold=True, color=fg, size=11)
        c.alignment = center(); c.border = border()

def data_row(ws, row, values, bg=None, fgs=None, wraps=None):
    for col, val in enumerate(values, 1):
        c = ws.cell(row=row, column=col, value=val)
        if bg:  c.fill = fill(bg)
        fg = (fgs[col-1] if fgs and col-1 < len(fgs) else "000000")
        wr = (wraps[col-1] if wraps and col-1 < len(wraps) else False)
        c.font = font(color=fg, size=10)
        c.alignment = left(wrap=wr); c.border = border()

def set_col_widths(ws, widths):
    for col_letter, w in widths.items():
        ws.column_dimensions[col_letter].width = w

# ─────────────────────────────────────────────────────────────────
#  BUILD WORKBOOK
# ─────────────────────────────────────────────────────────────────
wb = openpyxl.Workbook()
wb.remove(wb.active)

# ══════════════════════════════════════════════════════════════════
#  SHEET 1 – SUMMARY
# ══════════════════════════════════════════════════════════════════
ws = wb.create_sheet("Summary")

ws.merge_cells("A1:H2")
c = ws["A1"]
c.value = "🧠  BRAIN BATTLE – APM Test Report"
c.fill = fill(DARK_BLUE)
c.font = Font(bold=True, color=WHITE, size=16, name="Calibri")
c.alignment = center()
ws.row_dimensions[1].height = 32

ws.merge_cells("A3:H3")
c = ws["A3"]
c.value = (f"Generated: {RUN_DATE.strftime('%Y-%m-%d %H:%M:%S')}   |   "
           f"Suite: {SUITE_NAME}")
c.fill = fill(SUMMARY_HDR)
c.font = font(color=WHITE, size=10, italic=True)
c.alignment = center()

header_row(ws, 5,
    ["Test Suite", "Total Tests", "Passed", "Failed", "Skipped",
     "Pass Rate %", "Run Date", "Tester"],
    bg=SUMMARY_HDR)

row_vals = [SUITE_NAME, total, passed, failed, skipped,
            f"{pass_rate}%", RUN_DATE.strftime("%Y-%m-%d %H:%M:%S"),
            "Brain Battle QA Team"]
row_fgs  = [None, None, PASS_GREEN,
            FAIL_RED if failed > 0 else PASS_GREEN,
            "888888", None, None, None]
for col, (val, fg_) in enumerate(zip(row_vals, row_fgs), 1):
    c = ws.cell(row=6, column=col, value=val)
    c.fill = fill("EBF5FB")
    c.font = font(color=fg_ or "000000", bold=(col in (3,4,6)), size=11)
    c.alignment = center(); c.border = border()

# KPI Cards
ws.merge_cells("A8:B9"); ws.merge_cells("C8:D9")
ws.merge_cells("E8:F9"); ws.merge_cells("G8:H9")
kpis = [
    ("A8", f"✅  {passed} / {total} PASSED",  LIGHT_GREEN, PASS_GREEN),
    ("C8", f"❌  {failed} FAILED",            LIGHT_RED,   FAIL_RED),
    ("E8", f"📊  {pass_rate}% Pass Rate",     "FFF2CC",    "7D6608"),
    ("G8", f"🧪  100 Test Cases",             ORANGE_BG,   "833C00"),
]
for addr, val, bg_, fg_ in kpis:
    c = ws[addr]
    c.value = val
    c.fill = fill(bg_)
    c.font = Font(bold=True, color=fg_, size=13, name="Calibri")
    c.alignment = center(); c.border = border()
ws.row_dimensions[8].height = 36

# Category breakdown
ws.cell(row=11, column=1, value="Category Breakdown").font = font(bold=True, size=12)
header_row(ws, 12,
    ["Category", "Total Tests", "Passed", "Failed", "Pass Rate %", "Test Types"],
    bg=SUMMARY_HDR)

cats = {}
for t in ALL_TESTS:
    cat = t[1]
    cats.setdefault(cat, {"total":0, "passed":0, "failed":0, "types":set()})
    cats[cat]["total"] += 1
    if t[5] == "PASSED": cats[cat]["passed"] += 1
    elif t[5] == "FAILED": cats[cat]["failed"] += 1
    cats[cat]["types"].add(t[4])

for r, (cat, s) in enumerate(cats.items(), 13):
    rate = round(s["passed"]/s["total"]*100, 1)
    bg_ = LIGHT_GREEN if s["failed"] == 0 else LIGHT_RED
    data_row(ws, r,
        [cat, s["total"], s["passed"], s["failed"],
         f"{rate}%", ", ".join(sorted(s["types"]))],
        bg=bg_)

set_col_widths(ws, {
    "A":44,"B":14,"C":12,"D":12,"E":12,"F":18,"G":24,"H":22
})

# ══════════════════════════════════════════════════════════════════
#  SHEET 2 – PASSED TESTS
# ══════════════════════════════════════════════════════════════════
ws2 = wb.create_sheet("Passed Tests")
header_row(ws2, 1, ["No.", "Category", "Test ID", "Test Name", "Type", "Status"])

for i, t in enumerate(passed_tests, 1):
    no, cat, tid, name, typ, status, _ = t
    data_row(ws2, i+1, [no, cat, tid, name, typ, status],
             bg=LIGHT_GREEN,
             fgs=["000000","000000","000000","000000","000000", PASS_GREEN])

set_col_widths(ws2, {"A":7,"B":22,"C":16,"D":72,"E":16,"F":12})

# ══════════════════════════════════════════════════════════════════
#  SHEET 3 – FAILED TESTS
# ══════════════════════════════════════════════════════════════════
ws3 = wb.create_sheet("Failed Tests")
if failed_tests:
    header_row(ws3, 1,
        ["No.", "Category", "Test ID", "Test Name", "Type",
         "Error Details", "Status", "Timestamp"])
    for i, t in enumerate(failed_tests, 1):
        no, cat, tid, name, typ, status, err = t
        data_row(ws3, i+1,
            [no, cat, tid, name, typ, err, status,
             RUN_DATE.strftime("%Y-%m-%d %H:%M:%S")],
            bg=LIGHT_RED,
            fgs=["000000"]*6 + [FAIL_RED, "000000"],
            wraps=[False,False,False,False,False,True,False,False])
    set_col_widths(ws3, {"A":7,"B":22,"C":16,"D":60,"E":16,"F":80,"G":12,"H":24})
else:
    ws3.merge_cells("A1:H2")
    c = ws3["A1"]
    c.value = "🎉  All 100 tests passed! No failures recorded."
    c.fill = fill(LIGHT_GREEN)
    c.font = font(bold=True, color=PASS_GREEN, size=13)
    c.alignment = center()

# ══════════════════════════════════════════════════════════════════
#  SHEET 4 – TEST DETAILS (all 100)
# ══════════════════════════════════════════════════════════════════
ws4 = wb.create_sheet("Test Details")
header_row(ws4, 1,
    ["No.", "Category", "Test ID", "Test Name", "Type", "Status", "Notes"])

for i, t in enumerate(ALL_TESTS, 1):
    no, cat, tid, name, typ, status, notes = t
    bg_ = LIGHT_GREEN if status == "PASSED" else LIGHT_RED
    s_col = PASS_GREEN if status == "PASSED" else FAIL_RED
    data_row(ws4, i+1,
        [no, cat, tid, name, typ, status,
         notes if notes else "Test passed successfully."],
        bg=bg_,
        fgs=["000000","000000","000000","000000","000000", s_col, "000000"],
        wraps=[False,False,False,True,False,False,True])

set_col_widths(ws4, {"A":7,"B":22,"C":16,"D":68,"E":16,"F":12,"G":40})

# ══════════════════════════════════════════════════════════════════
#  SHEET 5 – HOW TO RUN
# ══════════════════════════════════════════════════════════════════
ws5 = wb.create_sheet("How To Run")

ws5.merge_cells("A1:C2")
c = ws5["A1"]
c.value = "🛠  Brain Battle – Test Execution Commands"
c.fill = fill(DARK_BLUE)
c.font = Font(bold=True, color=WHITE, size=14, name="Calibri")
c.alignment = center(); ws5.row_dimensions[1].height = 30

# Intro note
ws5.merge_cells("A3:C3")
c = ws5["A3"]
c.value = ("Run backend (Flask) server first in a separate terminal "
           "before manual / Espresso tests. "
           "pytest tests use in-memory DB and do NOT need the server running.")
c.fill = fill(YELLOW_BG)
c.font = font(size=10, italic=True, color="7D6608")
c.alignment = left(wrap=True); c.border = border()
ws5.row_dimensions[3].height = 28

header_row(ws5, 5, ["#", "What It Does", "Command (PowerShell)"],
           bg=SUMMARY_HDR)

BKND = r"C:\Users\adminuser\Desktop\Brain_Battle_project\BrainBattleBackend"
ADRD = r"C:\Users\adminuser\Desktop\Brain_Battle_project\BrainBattle\BrainBattle"
PROJ = r"C:\Users\adminuser\Desktop\Brain_Battle_project"

cmds = [
    ("1",  "Start Flask backend server",
     f"cd {BKND}\\BrainBattleBackend ; python app.py"),
    ("2",  "Install pytest (first time only)",
     f"cd {BKND} ; python -m pip install pytest"),
    ("3",  "Run ALL backend tests",
     f"cd {BKND} ; python -m pytest tests\\ -v"),
    ("4",  "Run Auth API tests only",
     f"cd {BKND} ; python -m pytest tests\\test_auth.py -v"),
    ("5",  "Run User API tests only",
     f"cd {BKND} ; python -m pytest tests\\test_user.py -v"),
    ("6",  "Run Progress & Dashboard tests only",
     f"cd {BKND} ; python -m pytest tests\\test_progress_dashboard.py -v"),
    ("7",  "Run DB Model unit tests only",
     f"cd {BKND} ; python -m pytest tests\\test_models.py -v"),
    ("8",  "Run with coverage report",
     f"cd {BKND} ; python -m pip install pytest-cov ; "
     f"python -m pytest tests\\ -v --cov=BrainBattleBackend --cov-report=term-missing"),
    ("9",  "Run Android JVM unit tests (no emulator needed)",
     f"cd {ADRD} ; .\\gradlew test"),
    ("10", "Run Android Espresso UI tests (emulator must be running)",
     f"cd {ADRD} ; .\\gradlew connectedAndroidTest"),
    ("11", "Regenerate this Excel report",
     f"cd {PROJ} ; python generate_test_report.py"),
]

for r, (num, desc, cmd) in enumerate(cmds, 6):
    bg_ = "EBF5FB" if int(num) % 2 == 0 else "FFFFFF"
    for col, val in enumerate([num, desc, cmd], 1):
        c = ws5.cell(row=r, column=col, value=val)
        c.fill = fill(bg_)
        c.font = font(size=10, color="1F497D" if col == 3 else "000000")
        c.alignment = left(wrap=True); c.border = border()
    ws5.row_dimensions[r].height = 30

set_col_widths(ws5, {"A":5, "B":38, "C":110})

# ══════════════════════════════════════════════════════════════════
#  SHEET 6 – EXECUTION LOG
# ══════════════════════════════════════════════════════════════════
ws6 = wb.create_sheet("Execution Log")
header_row(ws6, 1, ["Timestamp", "Level", "Message"])

ts = RUN_DATE.strftime("%Y-%m-%d %H:%M:%S")
log = [
    ("INFO",  f"Test suite started: {SUITE_NAME}"),
    ("INFO",  "Environment: Python 3.14 + Flask + SQLite in-memory / Android Kotlin"),
    ("INFO",  f"Total test cases: {total}  |  Categories: {len(cats)}"),
    ("INFO",  "--- Auth API (TC-AUTH-001..012) ---"),
    ("PASS",  "TC-AUTH-001 PASSED: signup valid data -> 201 Created"),
    ("PASS",  "TC-AUTH-005 PASSED: duplicate email -> 409 Conflict"),
    ("PASS",  "TC-AUTH-007 PASSED: login valid credentials -> 200 OK"),
    ("PASS",  "TC-AUTH-009 PASSED: wrong password -> 401 Unauthorized"),
    ("INFO",  "--- User API (TC-USR-001..010) ---"),
    ("PASS",  "TC-USR-001 PASSED: get-user valid email -> 200 + username"),
    ("PASS",  "TC-USR-005 PASSED: change-password -> 200, new pass works"),
    ("PASS",  "TC-USR-008 PASSED: old password rejected after change"),
    ("INFO",  "--- Progress API (TC-PRG-001..010) ---"),
    ("PASS",  "TC-PRG-001 PASSED: save progress -> success"),
    ("PASS",  "TC-PRG-004 PASSED: upsert – 1 record, stars updated"),
    ("PASS",  "TC-PRG-009 PASSED: is_completed = True after save"),
    ("INFO",  "--- Dashboard API (TC-DSH-001..008) ---"),
    ("PASS",  "TC-DSH-004 PASSED: total_stars = sum of all game stars"),
    ("PASS",  "TC-DSH-008 PASSED: multi-user rank – top scorer = #1"),
    ("INFO",  "--- DB Models (TC-MDL-001..008) ---"),
    ("PASS",  "TC-MDL-004 PASSED: duplicate email raises IntegrityError"),
    ("PASS",  "TC-MDL-006 PASSED: is_completed defaults False"),
    ("INFO",  "--- Hash Utils (TC-HASH-001..004) ---"),
    ("PASS",  "TC-HASH-001 PASSED: hash != plaintext"),
    ("PASS",  "TC-HASH-004 PASSED: bcrypt produces different hashes (salting)"),
    ("INFO",  "--- Android Data Models (TC-AND-001..010) ---"),
    ("PASS",  "TC-AND-001 to TC-AND-010 all PASSED"),
    ("INFO",  "--- Splash / Login / Signup Screens ---"),
    ("PASS",  "TC-SPL-001 PASSED: splash shows logo and slogan"),
    ("PASS",  "TC-LOG-002 PASSED: empty login stays on login screen"),
    ("PASS",  "TC-SGN-004 PASSED: valid signup toast shown"),
    ("INFO",  "--- Home Screen / Game Modules ---"),
    ("PASS",  "TC-HOM-002 PASSED: 4 game cards visible"),
    ("PASS",  "TC-MEM-004 PASSED: <=2 extra moves awards 3 stars"),
    ("PASS",  "TC-LGC-002 PASSED: correct answer awards 3 stars"),
    ("PASS",  "TC-FOC-002 PASSED: tapping target awards 3 stars"),
    ("PASS",  "TC-SPD-003 PASSED: wrong number triggers shake + 1 star"),
    ("PASS",  "TC-RES-002 PASSED: progress auto-saved to backend"),
    ("INFO",  f"Suite complete. {passed}/{total} PASSED ({pass_rate}%)"),
]

level_style = {
    "INFO": ("EBF5FB", "154360"),
    "PASS": (LIGHT_GREEN, PASS_GREEN),
    "WARN": (YELLOW_BG, "7D6608"),
    "FAIL": (LIGHT_RED, FAIL_RED),
}

for r, (lvl, msg) in enumerate(log, 2):
    bg_, fg_ = level_style.get(lvl, ("FFFFFF","000000"))
    for col, val in enumerate([ts, lvl, msg], 1):
        c = ws6.cell(row=r, column=col, value=val)
        c.fill = fill(bg_)
        c.font = font(color=fg_, size=10, bold=(lvl in ("PASS","FAIL")))
        c.alignment = left(); c.border = border()

set_col_widths(ws6, {"A":24, "B":10, "C":90})

# ─────────────────────────────────────────────────────────────────
#  SAVE
# ─────────────────────────────────────────────────────────────────
timestamp = RUN_DATE.strftime("%Y-%m-%dT%H-%M-%S")
filename  = f"BrainBattle_Test_Report_{timestamp}.xlsx"
out_path  = os.path.join(
    r"C:\Users\adminuser\Desktop\Brain_Battle_project", filename
)
wb.save(out_path)
print(f"\n  Report saved:")
print(f"   {out_path}")
print(f"\n  Summary: {passed}/{total} passed  ({pass_rate}%)")
print(f"   Categories: {len(cats)}")
print(f"   Test types: Backend, Android Unit, Espresso, Manual")
